#!/usr/bin/env python3
import csv
import json
import os
import shutil
import sys
from zipfile import ZipFile

import openpyxl
import pandas

# src
src_dir = "./dataset"
zip_file = "test.zip"
fn_csv_in_zip = 'arc/master/data/test.csv'

# dst
dst_dir = "./data/"


def main():
    print_header("1. Check version")
    print_version()

    print_header("2. Unpack zip")
    remove_dst()
    unpack_arc()

    print_header("3. Read csv file")
    # 3.1 Read csv file as list of list
    data_list = read_test_csv_list(dst_dir, fn_csv_in_zip)
    print("data_list:", data_list)
    print("data_list year:", data_list[1][0])
    # 3.2 Read csv file as list of dictionary
    data_dict = read_test_csv_dict(dst_dir, fn_csv_in_zip)
    print("data_dict:", data_dict)
    print("data_dict year:", data_dict[1]['year'])

    print_header("4. Convert csv to json and xlsx")
    data_json, filename_json = convert_csv_to_json(dst_dir, fn_csv_in_zip)
    print("json:", data_json)
    print("json filename:", filename_json)

    print_header("5. Convert json to xlsx")
    # 5.1 using lib openpyxl
    filename_xls1 = convert_json_to_xlsx_by_lib(filename_json)
    print(f"xls1: {filename_xls1}")
    # 5.2 using pandas
    filename_xls2 = convert_json_to_xlsx_by_pandas(filename_json)
    print(f"xls2: {filename_xls2}")


def print_header(msg):
    line = '=' * (len(msg) + 2)
    print(f'\n{line}')
    print(f' {msg} ')
    print(f'{line}')


def print_version():
    print(sys.version)


def remove_dst():
    print(f"Current workdir: {os.getcwd()}")
    if dst_dir and dst_dir not in ('/', './'):
        shutil.rmtree(dst_dir, ignore_errors=True)


def unpack_arc():
    print(sys.version)

    zip_full_path = os.path.join(src_dir, zip_file)
    with ZipFile(zip_full_path, 'r') as zip_obj:
        zip_obj.extractall(dst_dir)


def read_test_csv_list(dn, fn):
    filename = os.path.join(dn, fn)
    lines = []
    with open(filename, 'r') as data:
        for line in csv.reader(data):
            lines.append(line)

    return lines


def read_test_csv_dict(dn, fn):
    filename = os.path.join(dn, fn)
    with open(filename, 'r') as f:
        dict_reader = csv.DictReader(f)
        list_of_dict = list(dict_reader)
    return list_of_dict


def convert_csv_to_json(dn, fn_csv):
    data = read_test_csv_dict(dn, fn_csv)

    dn_json = os.path.dirname(fn_csv)
    fn_csv_name = os.path.basename(fn_csv)
    fn_parts = os.path.splitext(fn_csv_name)
    fn_name = fn_parts[0]
    filename_json = os.path.join(dn, dn_json, fn_name + '.json')

    # Serializing json
    data_json = json.dumps(data, indent=2)

    # Writing to json
    with open(filename_json, "w") as json_file:
        json_file.write(data_json)

    return data_json, filename_json


def convert_json_to_xlsx_by_lib(filename_json):
    fn_parts = os.path.splitext(filename_json)
    fn_name = fn_parts[0]
    filename_xlsx = fn_name + '_lib.xlsx'

    # Read json
    with open(filename_json, "r") as f:
        data = json.load(f)

    # Create a new workbook
    wb = openpyxl.Workbook()

    # Get the active worksheet
    ws = wb.active

    # Write the data to the worksheet
    for row_index, row_data in enumerate(data, 1):
        if row_index == 1:
            for col_index, col_data in enumerate(row_data.keys(), 1):
                ws.cell(row=row_index, column=col_index, value=col_data)

        for col_index, col_data in enumerate(row_data.values(), 1):
            ws.cell(row=row_index + 1, column=col_index, value=col_data)

    # Save the workbook to a file
    wb.save(filename_xlsx)

    return filename_xlsx


def convert_json_to_xlsx_by_pandas(filename_json):
    fn_parts = os.path.splitext(filename_json)
    fn_name = fn_parts[0]
    filename_xlsx = fn_name + '_pandas.xlsx'

    pandas.read_json(filename_json).to_excel(filename_xlsx, index=False)

    return filename_xlsx


if __name__ == '__main__':
    main()
