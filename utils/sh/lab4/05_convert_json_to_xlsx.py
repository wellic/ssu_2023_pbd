#!/usr/bin/env python3
import json
import os

import openpyxl
import pandas

# src
src_dir = "./dataset"
zip_file = "test.zip"
fn_csv_in_zip = 'arc/master/data/test.csv'

# dst
dst_dir = "./data/"


def main():
    print_header("5. Convert json to xlsx")
    # 5.1 using lib openpyxl
    filename_xls1 = convert_json_to_xlsx_by_lib(filename_json)
    print(f"xls1: {filename_xls1}")
    # 5.2 using pandas
    filename_xls2 = convert_json_to_xlsx_by_pandas(filename_json)
    print(f"xls2: {filename_xls2}")


def convert_json_to_xlsx_by_lib(filename_json):
    fn_parts = os.path.splitext(filename_json)
    fn_name = fn_parts[0]
    filename_xlsx = fn_name + '_lib.xlsx'

    # Read json
    with open(filename_json, "r") as f:
        data = json.load(f)

    # Create a new workbook
    wb = openpyxl.Workbook()

    # Get the active worksheet
    ws = wb.active

    # Write the data to the worksheet
    for row_index, row_data in enumerate(data, 1):
        if row_index == 1:
            for col_index, col_data in enumerate(row_data.keys(), 1):
                ws.cell(row=row_index, column=col_index, value=col_data)

        for col_index, col_data in enumerate(row_data.values(), 1):
            ws.cell(row=row_index + 1, column=col_index, value=col_data)

    # Save the workbook to a file
    wb.save(filename_xlsx)

    return filename_xlsx


def convert_json_to_xlsx_by_pandas(filename_json):
    fn_parts = os.path.splitext(filename_json)
    fn_name = fn_parts[0]
    filename_xlsx = fn_name + '_pandas.xlsx'

    pandas.read_json(filename_json).to_excel(filename_xlsx, index=False)

    return filename_xlsx
